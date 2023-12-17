import json

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class Converter():
    def __init__(
        self,
        file_path: str,
        save_name: str,
        rows_per_line: int = None
    ) -> None:
        self.file_path = file_path
        self.save_name = save_name
        self.json_data = self.convert_json_to_dict(self.file_path)
        self.rows_per_line = rows_per_line or self.calculate_rows_per_line()

    @staticmethod
    def convert_json_to_dict(file_path: str):
        with open(file_path, 'r') as json_file:
            data = json.load(json_file)
        return data

    def calculate_rows_per_line(self):

        max_lines: int = 0

        for main_line in self.json_data:
            for line in main_line['lines']:
                if len(line['moves']) > max_lines:
                    max_lines = len(line['moves'])

        return max_lines

    def create_excel_file(self):

        wb = Workbook()
        ws = wb.active

        # Initial settings for the layout
        starting_row = 1
        max_rows_in_line = 0

        for main_line in self.json_data:
            title = main_line['title']

            for line_index, line in enumerate(main_line['lines']):
                line_name = line['name']
                chess_moves = []
                current_line_depth = 0
                for move in line['moves']:
                    current_line_depth += 1
                    chess_moves.append(
                        [
                            move['move_number'],
                            move['white'],
                            move['black']
                        ]
                    )

                if self.rows_per_line > current_line_depth:
                    for _ in range(self.rows_per_line - current_line_depth):
                        chess_moves.append(['', '', ''])

                df = pd.DataFrame(
                    chess_moves,
                    columns=['Move Number', 'White', 'Black']
                )

                # Calculate the starting column (A or E) based on the line index
                starting_col = 'A' if line_index % 2 == 0 else 'E'

                # Calculate the starting row for new lines
                if line_index % 2 == 0:
                    starting_row += max_rows_in_line + 2
                    max_rows_in_line = 0  # Reset for the new pair of lines

                # Set title and line name with formatting
                title_cell = ws[f'{starting_col}{starting_row + 1}']
                title_cell.value = title
                title_cell.font = Font(size=9)

                line_name_cell = ws[f'{starting_col}{starting_row + 2}']
                line_name_cell.value = line_name
                line_name_cell.font = Font(bold=True)

                # Define a border style
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # Add rows to the worksheet from the DataFrame
                for r_idx, row in enumerate(
                    dataframe_to_rows(
                        df,
                        index=False,
                        header=True
                    ), starting_row + 3
                ):
                    for c_idx, value in enumerate(row, 1):
                        col = chr(ord(starting_col) + c_idx - 1)
                        cell = ws[f'{col}{r_idx}']
                        cell.value = value
                        if r_idx == starting_row + 3:  # This is the header row
                            cell.font = Font(bold=True)
                        cell.border = thin_border

                # Keep track of the number of rows used in the longest line
                max_rows_in_line = max(max_rows_in_line, len(chess_moves) + 2)

        # Save the workbook to a file
        excel_path = f'{self.save_name}.xlsx'
        wb.save(excel_path)

        print(f"The Excel file with multiple lines has been saved to {excel_path}")


    def create_box_around_lines(self):
        pass


if __name__ == '__main__':
    converter = Converter(
        file_path='files/advance_caro_chessly_lines.json',
        save_name='advance_caro_kann_chessly_lines'
    )
    converter.create_excel_file()
